#!/usr/bin/env python3
"""
Vérifie, pour une liste de codeap donnés, s'il existe des flux sur le cluster de
logs, et liste les datasets associés. Sortie : un fichier Excel, une ligne = un codeap.

Entrées :
  --codeaps : fichier texte listant les codeap à vérifier (un par ligne)
              accepte les formats 'ap12345', 'a12345', ou juste '12345'
  --indices : sortie de  GET _cat/indices/logs-*?h=index  (noms d'index, colonnes en trop ignorées)
  --output  : chemin du fichier .xlsx à produire (défaut: codeap_flux.xlsx)

Usage :
  python3 codeap_flux.py --codeaps codeaps.txt --indices indices.txt --output resultat.xlsx

Convention de nommage d'index :
  logs-<codename>.<dataset>-<codeap>_<env>_<retention>_<date>
"""

import sys
import re
import argparse
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Extraction (codename, dataset, codeap, env) depuis un nom d'index ---
FLUX_PATTERN = re.compile(
    r"^logs-"
    r"(?P<codename>[^.]+)\."
    r"(?P<dataset>.+)"
    r"-(?P<codeap>ap?\d+)"
    r"_(?P<env>[a-zA-Z0-9]+)"
    r"_(?P<retention>r\d+)"
    r"_(?P<date>.+)$"
)


def normalize_codeap(raw: str) -> str:
    """
    Normalise un codeap vers sa forme canonique en ne gardant que les chiffres.
    'ap12345' -> '12345', 'a12345' -> '12345', '12345' -> '12345'.
    Permet de matcher indépendamment du préfixe (a / ap) utilisé dans les index.
    """
    digits = re.sub(r"\D", "", raw.strip())
    return digits


def extract_index_name(line: str):
    """Récupère le nom d'index d'une ligne (gère la sortie _cat avec colonnes)."""
    for tok in line.split():
        if tok.startswith("logs-"):
            return tok
    return None


def main():
    parser = argparse.ArgumentParser(description="Vérifie la présence de flux par codeap.")
    parser.add_argument("--codeaps", required=True, help="Fichier des codeap à vérifier (un par ligne)")
    parser.add_argument("--indices", required=True, help="Fichier des noms d'index (_cat/indices)")
    parser.add_argument("--output", default="codeap_flux.xlsx", help="Fichier Excel de sortie")
    args = parser.parse_args()

    # 1. Construire l'inventaire du cluster : codeap normalisé -> infos flux
    #    On agrège par (dataset) et on garde codenames + environnements rencontrés.
    inventory = defaultdict(lambda: {
        "datasets": set(),
        "codenames": set(),
        "envs": set(),
        "index_count": 0,
    })

    with open(args.indices, "r", encoding="utf-8") as f:
        for line in f:
            idx = extract_index_name(line)
            if idx is None:
                continue
            m = FLUX_PATTERN.match(idx)
            if not m:
                continue
            key = normalize_codeap(m.group("codeap"))
            entry = inventory[key]
            entry["datasets"].add(m.group("dataset"))
            entry["codenames"].add(m.group("codename"))
            entry["envs"].add(m.group("env"))
            entry["index_count"] += 1

    # 2. Lire la liste des codeap à vérifier (on conserve la forme d'origine pour l'affichage)
    codeaps_input = []
    with open(args.codeaps, "r", encoding="utf-8") as f:
        for line in f:
            val = line.strip()
            if val:
                codeaps_input.append(val)

    # 3. Construire le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Flux par codeap"

    headers = ["Codeap", "Flux présent ?", "Nb datasets", "Datasets", "Codenames", "Environnements", "Nb index"]
    ws.append(headers)

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E78")
    yes_fill = PatternFill("solid", start_color="C6EFCE")   # vert clair
    yes_font = Font(name="Arial", color="006100", bold=True)
    no_fill = PatternFill("solid", start_color="FFC7CE")    # rouge clair
    no_font = Font(name="Arial", color="9C0006", bold=True)
    normal_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    # 4. Une ligne par codeap
    for raw_codeap in codeaps_input:
        key = normalize_codeap(raw_codeap)
        entry = inventory.get(key)
        present = entry is not None and entry["index_count"] > 0

        if present:
            datasets = sorted(entry["datasets"])
            codenames = sorted(entry["codenames"])
            envs = sorted(entry["envs"])
            row = [
                raw_codeap,
                "OUI",
                len(datasets),
                ", ".join(datasets),
                ", ".join(codenames),
                ", ".join(envs),
                entry["index_count"],
            ]
        else:
            row = [raw_codeap, "NON", 0, "", "", "", 0]

        ws.append(row)
        r = ws.max_row

        # Style de la ligne
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = border
            if col_idx == 1:
                cell.font = Font(name="Arial", bold=True, size=10)
                cell.alignment = center
            elif col_idx == 2:
                cell.alignment = center
                if present:
                    cell.fill = yes_fill
                    cell.font = yes_font
                else:
                    cell.fill = no_fill
                    cell.font = no_font
            elif col_idx in (3, 7):
                cell.font = normal_font
                cell.alignment = center
            else:
                cell.font = normal_font
                cell.alignment = left_wrap

    # 5. Largeurs de colonnes
    widths = {"A": 14, "B": 14, "C": 12, "D": 45, "E": 30, "F": 18, "G": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Fige la ligne d'en-tête + filtre auto
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    wb.save(args.output)

    # Résumé console
    n_present = sum(
        1 for rc in codeaps_input
        if inventory.get(normalize_codeap(rc)) and inventory[normalize_codeap(rc)]["index_count"] > 0
    )
    print(f"Codeap vérifiés : {len(codeaps_input)}")
    print(f"  avec flux     : {n_present}")
    print(f"  sans flux     : {len(codeaps_input) - n_present}")
    print(f"Fichier généré  : {args.output}")


if __name__ == "__main__":
    main()
